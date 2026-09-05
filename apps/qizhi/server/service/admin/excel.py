"""
openpyxl 通用包装：构建带表头样式与列宽的工作簿。
"""

from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def build_workbook(sheet_name: str, headers: Sequence[str], rows: Sequence[Sequence]) -> bytes:
    """生成 xlsx 字节内容。

    表头：加粗居中 + 浅蓝底色（与 HomeView agent badge 同色调）。
    列宽：依据内容长度自适应，cap 60 字符。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(list(headers))

    header_fill = PatternFill("solid", fgColor="DEE8FF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    for row in rows:
        ws.append(list(row))

    # 列宽自适应（中文字符按 2 估算，但 openpyxl 不区分；取保守上限）
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            value = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            text = "" if value is None else str(value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
